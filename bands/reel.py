import pandas as pd
import os
from openpyxl import Workbook

def reel(file_path, output_path="output.xlsx"):
    # 判断文件是否存在
    if not os.path.exists(file_path):
        print('文件不存在')
        return

    # # 根据文件后缀名读取文件
    # if file_path.endswith('.xlsx'):
    #     df = pd.read_excel(file_path, index_col=None)
    # elif file_path.endswith('.csv'):
    #     df = pd.read_csv(file_path, index_col=None)
    # else:
    #     print('不支持的文件格式')
    #     return

    df = pd.ExcelFile(file_path)

    # 获取文件所在的目录
    file_dir = os.path.abspath(os.path.dirname(file_path))

    # 创建 Excel 工作簿
    excel_path = os.path.join(file_dir, output_path)
    workbook = Workbook()

    # 创建结果工作表
    worksheet = workbook.create_sheet('Result')

    # 添加表头 [machine_name, column, icon, weight, index]
    worksheet.append(['machine_name', 'column', 'icon', 'weight', 'index'])

    # 读取每个工作表
    for sheet_name in df.sheet_names:
        # 表名为 machine_name
        machine_name = sheet_name

        # 获取当前工作表的数据
        sheet_data = df.parse(sheet_name)

        # 遍历每一列（每两列为一组：icon 和 weight）
        for col_index in range(0, len(sheet_data.columns), 2):
            # 获取当前 column 的 icon 和 weight 列
            icon_col = sheet_data.iloc[:, col_index]
            weight_col = sheet_data.iloc[:, col_index + 1]

            # 遍历每一行
            for row_index in range(len(icon_col)):
                # 如果 icon 为 NaN，则跳过
                if pd.isna(icon_col.iloc[row_index]):
                    continue
                # 添加数据
                worksheet.append([
                    machine_name,  # machine_name
                    col_index // 2,  # column (0, 1, 2, ...)
                    icon_col.iloc[row_index],  # icon
                    weight_col.iloc[row_index],  # weight
                    row_index  # index
                ])

    # 删除默认创建的 Sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # 保存 Excel 文件
    workbook.save(excel_path)
    print(f'转换完成，结果已保存到 {excel_path}')

# 调用函数
reel('425/reel_425.xlsx')